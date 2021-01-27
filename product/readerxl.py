import os
import openpyxl
from django.db.backends import sqlite3
from product.models import Product


def readxl(file_xlx):
    try:
        wb_obj = openpyxl.load_workbook(file_xlx)
        sheet = wb_obj.active

        total_products = 0

        for row in sheet.iter_rows(max_row=sheet.max_row):
            if not (row[0].value is None) and type(row[0].value) == int:
                # print('in if1 TRUE conditions')
                article = row[0].value
                category = row[1].value
                company = row[2].value
                model = row[3].value
                start_price = row[4].value
                end_price = row[5].value
                defaults = {'title': category,
                            'price': start_price,
                            'quantity': 5,
                            'is_active': True,
                            'description': model}
                instance, created = Product.objects.update_or_create(id=article, defaults=defaults)
                # print('create/update product:', created)
                # print('instance:', instance)
                total_products += 1
            else:
                # print('in else1 condition')
        print('Total:', total_products)
    except:
        # print('Raise error')
